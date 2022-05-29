import pandas as pd
import openpyxl
import wx
import wx.grid as gridlib
import numpy as np

def GetSheetNameListFromExcelFileName(fileName):
    wb = openpyxl.load_workbook(fileName)
    sheets = wb.worksheets
    result = []
    for sheet in sheets:
        result.append(sheet.title)
    return result

def GetSheetDataFromExcelFileName(fileName,sheetName):
    wb = openpyxl.load_workbook(fileName)
    ws = wb.get_sheet_by_name(sheetName)
    data = []
    for row in ws.values:
        temp=[]
        for value in row:
            temp.append(value)
        data.append(temp)
    data = np.array(data)
    return data

def ModifySurfaceXConfigurationExcelFile(leftEnable,leftBendValue,rightEnable,rightBendValue,topEnable,topBendValue,bottomEnable,bottomBendValue,bottomCutEnable,bottomBendCutValue):
    wb = openpyxl.load_workbook("D:\\WorkSpace\\Solidworks\\N.2SA\\SurfaceXSLDPRT.xlsx")
    ws = wb.get_sheet_by_name("Sheet1")
    ws['G3'] = 'U' if leftEnable else "S"
    ws['H3'] = 'U' if rightEnable else "S"
    ws['E3'] = 'U' if bottomEnable else "S"
    ws['F3'] = 'U' if bottomCutEnable else "S"
    ws['I3'] = bottomBendCutValue if bottomCutEnable and bottomCutEnable else 0
    ws['J3'] = bottomBendValue if bottomCutEnable else 0
    ws['K3'] = leftBendValue if leftEnable else 0
    ws['L3'] = rightBendValue if rightEnable else 0
    ws['M3'] = topBendValue if rightEnable else 0
    ws['N3'] = 'U' if topEnable else "S"

    wb.save('D:\\WorkSpace\\Solidworks\\N.2SA\\SurfaceXSLDPRT.xlsx')

def ModifySurfaceYConfigurationExcelFile(leftBendEnable, leftBendValue, rightBendEnable, rightBendValue,
                                     topBendEnable, topBendValue, bottomBendEnable, bottomBendValue,
                                     bottomBendCutEnable, bottomBendCutValue, bottomRaiseEnable,
                                     bottomRaiseValue,leftSelvedgeEnable,rightSelvedgeEnable):
    wb = openpyxl.load_workbook("D:\\WorkSpace\\Solidworks\\N.2SA\\SurfaceYSLDPRT.xlsx")
    ws = wb.get_sheet_by_name("Sheet1")
    ws['D3'] = 'U' if leftBendEnable else "S"
    ws['E3'] = 'U' if rightBendEnable else "S"
    ws['F3'] = 'U' if topBendEnable else "S"
    ws['G3'] = 'U' if bottomBendEnable else "S"
    ws['H3'] = 'U' if bottomBendCutEnable else "S"
    ws['I3'] = 'U' if bottomRaiseEnable else "S"
    ws['J3'] = 'U' if leftSelvedgeEnable else "S"
    ws['K3'] = 'U' if rightSelvedgeEnable else "S"
    ws['L3'] = leftBendValue if leftBendEnable else 0
    ws['M3'] = rightBendValue if rightBendEnable else 0
    ws['N3'] = topBendValue if topBendEnable else 0
    ws['O3'] = bottomBendValue if bottomBendEnable else 0
    ws['P3'] = bottomBendCutValue if bottomBendEnable and bottomBendCutEnable else 0
    wb.save('D:\\WorkSpace\\Solidworks\\N.2SA\\SurfaceYSLDPRT.xlsx')


def ModifyRockWoolConfigurationExcelFile(leftSlotEnable,leftSlotDepthValue,rightSlotEnable,rightSlotDepthValue,topSlotEnable,topSlotDepthValue,bottomSlotEnable,bottomSlotDepthValue,bottomRaiseEnable,bottomRaiseValue):
    wb = openpyxl.load_workbook("D:\\WorkSpace\\Solidworks\\N.2SA\\岩棉配置表.xlsx")
    ws = wb.get_sheet_by_name("Sheet1")
    ws['D3'] = 'U' if leftSlotEnable else "S"
    ws['E3'] = 'U' if rightSlotEnable else "S"
    ws['F3'] = 'U' if topSlotEnable else "S"
    ws['G3'] = 'U' if bottomSlotEnable else "S"
    ws['H3'] = 'U' if bottomSlotEnable else "S"
    ws['I3'] = leftSlotDepthValue if leftSlotEnable else 0
    ws['J3'] = rightSlotDepthValue if rightSlotEnable else 0
    ws['K3'] = topSlotDepthValue if topSlotEnable else 0
    ws['L3'] = bottomSlotDepthValue if bottomSlotEnable else 0
    ws['M3'] = bottomRaiseValue if bottomRaiseEnable else 0
    wb.save('D:\\WorkSpace\\Solidworks\\N.2SA\\岩棉配置表.xlsx')
